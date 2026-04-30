"""Excel change report."""

from __future__ import annotations

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

HEADER_FILL = PatternFill("solid", fgColor="305496")
HEADER_FONT = Font(bold=True, color="FFFFFF")
PASS_FILL   = PatternFill("solid", fgColor="C6EFCE")
FAIL_FILL   = PatternFill("solid", fgColor="FFC7CE")
MONO_FONT   = Font(name="Consolas", size=10)


def write(results: list[dict],
          validations: list[dict],
          out_dir: str | Path,
          mode: str) -> Path:
    out_dir = Path(out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    path = out_dir / f"change_report_{mode}_{ts}.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    headers = ["Host", "Mode", "OK", "Validation", "Findings"]
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")

    val_by_host = {v["host"]: v for v in validations}
    for r in sorted(results, key=lambda x: x["host"]):
        v = val_by_host.get(r["host"])
        passed = v["passed"] if v else None
        findings = " | ".join(v["findings"]) if v and v["findings"] else ""
        ws.append([
            r["host"], r["mode"], "yes" if r.get("ok") else "no",
            ("PASS" if passed else "FAIL") if v else "n/a",
            findings,
        ])
        row = ws.max_row
        if v is not None:
            ws.cell(row=row, column=4).fill = PASS_FILL if passed else FAIL_FILL

    for col_letter, width in zip("ABCDE", (22, 12, 6, 12, 80)):
        ws.column_dimensions[col_letter].width = width

    for r in results:
        if not r["diff"]:
            continue
        sheet = wb.create_sheet(title=_safe_sheet_title(r["host"]))
        sheet.append(["config diff (running -> intended)"])
        sheet.cell(row=1, column=1).font = HEADER_FONT
        sheet.cell(row=1, column=1).fill = HEADER_FILL
        for line in r["diff"][:1000]:
            sheet.append([line])
            sheet.cell(row=sheet.max_row, column=1).font = MONO_FONT
        sheet.column_dimensions["A"].width = 110

    wb.save(path)
    return path


def _safe_sheet_title(name: str) -> str:
    bad = '[]:*?/\\'
    cleaned = "".join("_" if ch in bad else ch for ch in name)
    return cleaned[:31]
